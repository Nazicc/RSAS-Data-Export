import os
import re
import sys
import html
import queue
import shutil
import zipfile
import datetime
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class Vul_re(object):
    def __init__(self):
        super(Vul_re, self).__init__()
        self.vul_list_re = '<python>ip<python>.*?<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<td>(.*?)</td>.*?(\d+-\d+-\d+\s\d+:\d+:\d+).*?</td>.*?(\d+-\d+-\d+\s\d+:\d+:\d+).*?</td>.*?<python>host</python>.*?<python>vul_list<python>(.*?)<python>vul_list</python>.*?<python>ip</python>'
        self.vul_ip_re = '(<python>ip<python>.*?<python>ip</python>)'
        self.vul_detail_re = '<python>vul_detail<python>(.*?)<python>vul_detail</python>'
        self.vul_details_re = '<python>vul_details<python>(.*?)<python>vul_details</python>'

        self.danger_re = '<span class="level_danger_(.*?)".*?table_\d_(\d+).*?>(.*?)</span>'
        self.title_re = '<python>title<python>(.*?)<python>title</python>'
        self.time_re = '<python>host<python>.*?(\d+-\d+-\d+).*?<python>host</python>'
        self.scan_time_re = '<python>host<python>.*?(\d+-\d+-\d+\s\d+:\d+:\d+).*?<python>host</python>'
        self.other_re = '<td class="vul_port">(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>.*?<ul>(.*?)</ul>'
        self.host_name = '<python>host<python>.*?<td valign="top".*?<th>主机名</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>'

class File_re(object):
    def __init__(self):
        super(File_re, self).__init__()
        self.file_re = '.*?.zip'
        self.uzip_re = '.*?.html'
        self.all_title_re = '<th width="120">任务名称</th>.*?<td>(.*?)</td>'
        self.vul_list_re = '(<table id="vuln_list" class="report_table">.*?</table>)'
        self.vul_detail_re = '(<div id="vul_detail">.*?</div>)'
        self.vul_details_re = '(<tr class="solution.*?">.*?<td>.*?<table class="report_table plumb".*?>.*?</table>.*?</td>.*?</tr>)'
        self.host_re = '(<td valign="top" style="width:50%;">.*?<table class="report_table plumb">.*?<tbody>.*?<th width="120">IP地址</th>.*?</tbody>.*?</table></td>)'

class Vul_content(object):
    def __init__(self,vul_re):
        super(Vul_content, self).__init__()
        self.vul_ip_content = re.findall(vul_re.vul_ip_re,htmlcont,re.S|re.M)
        self.vul_detail_content = re.findall(vul_re.vul_detail_re,htmlcont,re.S|re.M)

class Solve_re(object):
    def __init__(self):
        super(Solve_re, self).__init__()
        self.solve_re = '<th width="100">解决办法</th>.*?<td>(.*?)</td>'
        self.describe_re = '<tr class="solution.*?table_\d_(\d+).*?<th width="100">详细描述</th>.*?<td>(.*?)</td>'
        self.cve_re = '<th width="100">CVE编号</th>.*?<td><a target=.*?>(.*?)</a>.*?</td>'

class Other(object):
    def __init__(self, vul_re, all_vuln_list):
        super(Other, self).__init__()
        self.all_other = re.findall(vul_re.other_re,all_vuln_list,re.S|re.M)

class Danger(object):
    def __init__(self, vul_re, other):
        super(Danger, self).__init__()
        self.danger_coneent = re.findall(vul_re.danger_re,other,re.S|re.M)

class Solve(object):
    def __init__(self, solve, all_vul_details):
        super(Solve, self).__init__()
        self.solve_plumb = re.findall(solve.solve_re,all_vul_details,re.S|re.M)
        self.describe_plumb = re.findall(solve.describe_re,all_vul_details,re.S|re.M)
        self.cve_plumb = re.findall(solve.cve_re,all_vul_details,re.S|re.M)

class Port_File_re(object):
    def __init__(self):
        super(Port_File_re, self).__init__()
        self.file_re = '.*?.zip'
        self.uzip_re = '.*?.html'
        self.all_title_re = '<th width="120">任务名称</th>.*?<td>(.*?)</td>'
        self.host_re = '<th width="120">IP地址</th>.*?<td>(\d+.\d+.\d+.\d+)</td>.*?<th>扫描起始时间</th>.*?<td>(\d+-\d+-\d+).*?</td>.*?<thead>.*?<th>端口</th>.*?<th>协议</th>.*?<th>服务</th>.*?<th>状态</th>.*?</thead>.*?<tbody>(.*?)</tbody>'
        self.port_re = '<tr class=".*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?</tr>'
        self.http_re = '.*?http.*?'
        self.https_re = '.*?https.*?'
        self.www_re = '.*?www.*?'

class WorkThread(QThread):
    log_return = pyqtSignal(str)
    def __init__(self,input_Button_cent,output_Button_cent,hight_status,middle_status,low_status,port_status,web_status):
        super(WorkThread, self).__init__()
        self.folder_start = input_Button_cent
        self.folder_end = output_Button_cent
        self.hight_status = hight_status
        self.middle_status = middle_status
        self.low_status = low_status
        self.port_status = port_status
        self.web_status = web_status
        
    def run(self):
        set_vul = []
        excel = {}
        letter = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

        with open('set.ini') as set_ini:
            ini = set_ini.readlines()[1:3]
            self.name_ini = ini[1].split('=')[1].strip()
            self.company_ini = ini[0].split('=')[1].strip()

        try:
            with open('set.ini') as vulnerable_ini:
                for vul in vulnerable_ini.readlines()[11:]:
                    set_vul.append(vul.strip().split('|'))
        except Exception as e:
            pass
        
        with open('set.ini') as cent:
            len_excel = 2
            for ini,column in zip(cent.readlines()[5].split('|'),letter):
                sign = ini.split(':')
                excel[column] = sign[0],sign[1],sign[2].strip()
                len_excel += 1

        try:
            shutil.rmtree('temp')
        except Exception as e:
            pass

        if self.hight_status or self.middle_status or self.low_status == True:
            try:
                shutil.rmtree(self.folder_end+'/汇总-漏洞跟踪表')
            except Exception as e:
                pass
            starttime = datetime.datetime.now()
            self.log_return.emit('正在提取数据...')

            os.mkdir('temp')
            with open('temp/database.mdb', 'w',encoding='gb18030') as content:
                content.write('')

            self.dirList = os.listdir(self.folder_start)
            for name in self.dirList:
                all_file_name = re.findall(File_re().file_re,name)
                for file_name in all_file_name:
                    try:
                        uzip = zipfile.ZipFile(self.folder_start+'/'+file_name)
                    except Exception as e:
                        self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                        self.log_textEdit.insertPlainText('\n{}不是正确的ZIP压缩包，请检查！\n\n'.format(file_name))
                        QtWidgets.QApplication.processEvents()
                        return e
                    try:
                        for uzip_content in uzip.namelist():
                            all_uzip_content = re.findall(File_re().uzip_re,uzip_content)
                            for all_uzip in all_uzip_content:
                                htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                                title = re.findall(File_re().all_title_re,htmlcont_zip,re.S|re.M)
                                for title_content in title:
                                    with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                        content.write('<python>title<python>')
                                        content.write(html.unescape(title_content))
                                        content.write('<python>title</python>\n')

                                    with open('temp/database.mdb','a',encoding='gb18030') as content:
                                        content.write('temp/'+html.unescape(title_content)+'.mdb\n')

                                host = re.findall(File_re().host_re,htmlcont_zip,re.S|re.M)
                                for host_content in host:
                                    with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                        content.write('<python>ip<python>\n')
                                        content.write('<python>host<python>\n')
                                        content.write(html.unescape(host_content))
                                        content.write('\n<python>host</python>\n')

                                vul_list = re.findall(File_re().vul_list_re,htmlcont_zip,re.S|re.M)
                                for list_content in vul_list:
                                    with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                        content.write('<python>vul_list<python>\n')
                                        content.write(html.unescape(list_content))
                                        content.write('\n<python>vul_list</python>\n')

                                vul_detail = re.findall(File_re().vul_detail_re,htmlcont_zip,re.S|re.M)
                                for detail_content in vul_detail:
                                    with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                        content.write('<python>vul_detail<python>\n')

                                    vul_details = re.findall(File_re().vul_details_re,detail_content,re.S|re.M)
                                    for list_details in vul_details:
                                        with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                            content.write('<python>vul_details<python>\n')
                                            content.write(html.unescape(list_details))
                                            content.write('\n<python>vul_details</python>\n')

                                with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                    content.write('\n<python>vul_detail</python>\n')
                                    content.write('<python>ip</python>\n')

                        self.log_return.emit('提取 {}'.format(title_content))
                    except Exception as e:
                        self.log_return.emit('{} 已被编辑，保存编码不一样，导出失败！'.format(file_name))
                        return e

            self.log_return.emit('\n数据提取完成，正在生成漏洞跟踪表...')
            os.mkdir(self.folder_end+'/汇总-漏洞跟踪表')

            vul_re = Vul_re()
            i = 1
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.freeze_panes = 'A2'
            ws.title = '系统漏洞'
            ws.column_dimensions['A'].width = 6
            ws['A1'] = '序号'
            for longs,column in zip(range(2,len(excel)+2),letter):
                ws.column_dimensions[column].width = excel[column][1]
                ws['{}1'.format(column)] = excel[column][2]

            # 样式
            font = Font(size=10, name='宋体')
            thin = Side(border_style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # 对齐
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            with open('set.ini') as colour:
                for color in colour.readlines()[8].strip().split('|'):
                    one_color = color.split(':')
                    lter = one_color[0].split('-')
                    #设置第一行的格式
                    title_font = Font(size=12, bold=True, name='宋体', color= one_color[1])
                    ws.row_dimensions[1].height = 40
                    for title_style in ws['{}1:{}1'.format(lter[0],lter[1])]:
                        for title_cell in title_style:
                            title_cell.font = title_font
                            title_cell.border = border
                            title_cell.alignment = alignment

            with open('temp/database.mdb',encoding='gb18030') as content:
                for zip_content in content:
                    vul_all_list = []
                    vul_all_detail = {}

                    zip_cont = zip_content.strip()
                    content = open(zip_cont,'r',encoding='gb18030')
                    global htmlcont
                    htmlcont = content.read()
                    content.close()
                    
                    sheet_name =  re.findall(vul_re.title_re,htmlcont,re.S|re.M)[0]
                    sheet_time =  re.findall(vul_re.time_re,htmlcont,re.S|re.M)[0]
                    vul_time = sheet_time.split('-')[1]
                    if vul_time[0] == '0':
                        vul_time = vul_time[1]
                    vul_time = vul_time + '月'
                    self.log_return.emit('正在导出 {}'.format(sheet_name))

                    vul_content = Vul_content(vul_re)

                    for all_vul_ip in vul_content.vul_ip_content:
                        vul_list_content = re.findall(vul_re.vul_list_re,all_vul_ip,re.S|re.M)
                        host_name = re.findall(vul_re.host_name,all_vul_ip,re.S|re.M)
                        if host_name:
                            pass
                        else:
                            host_name = ['']
                        for all_vul_list in vul_list_content:
                            for other in Other(vul_re,all_vul_list[3]).all_other:
                                for danger in Danger(vul_re,other[3]).danger_coneent:
                                    vul_name = danger[0]
                                    if len(set_vul) > 0:
                                        for set_vuls in set_vul:
                                            if set_vuls[1] == danger[2]:
                                                vul_name = set_vuls[0]
                                                break
                                    vul_all_list.append([danger[1],all_vul_list[0],danger[2].strip(),vul_name.replace('low','低').replace('middle','中').replace('high','高'),other[0],other[1],other[2],all_vul_list[1],all_vul_list[2],host_name[0]])

                    for all_vul_detail in vul_content.vul_detail_content:
                        vul_details_content = re.findall(vul_re.vul_details_re,all_vul_detail,re.S|re.M)
                        for all_vul_details in vul_details_content:
                            vul_detail = Solve(Solve_re(),all_vul_details)
                            for solve,describe in zip(vul_detail.solve_plumb,vul_detail.describe_plumb):
                                cve = vul_detail.cve_plumb
                                if cve:
                                    pass
                                else:
                                    cve = ['漏洞暂无CVE编号']
                                vul_all_detail[describe[0]] = re.sub('\s{2,}','\n',html.unescape(re.sub('\s{2,}','',solve)).replace('<br/>','\n')),re.sub('\s{2,}','\n',html.unescape(re.sub('\s{2,}','',describe[1])).replace('<br/>','\n')),cve[0]


                    vul_amount = []

                    for wait_list in vul_all_list:
                        wait_detail = vul_all_detail[wait_list[0]]
                        vul_amount.append(['',self.company_ini,sheet_name,wait_list[9],wait_list[1],wait_list[4],wait_list[5],wait_list[6],wait_list[2],'漏洞',wait_list[3],wait_detail[0].strip(),wait_detail[1],wait_detail[2],wait_list[7],wait_list[8],vul_time])

                    for amount in vul_amount:
                        if self.hight_status == True and amount[10] == '高':
                            ws.row_dimensions[i+1].height = 25
                            ws.append([i,amount[int(excel['B'][0]) if 'B' in excel else 0],amount[int(excel['C'][0]) if 'C' in excel else 0],amount[int(excel['D'][0]) if 'D' in excel else 0],amount[int(excel['E'][0]) if 'E' in excel else 0],amount[int(excel['F'][0]) if 'F' in excel else 0],amount[int(excel['G'][0]) if 'G' in excel else 0],amount[int(excel['H'][0]) if 'H' in excel else 0],amount[int(excel['I'][0]) if 'I' in excel else 0],amount[int(excel['J'][0]) if 'J' in excel else 0],amount[int(excel['K'][0]) if 'K' in excel else 0],amount[int(excel['L'][0]) if 'L' in excel else 0],amount[int(excel['M'][0]) if 'M' in excel else 0],amount[int(excel['N'][0]) if 'N' in excel else 0],amount[int(excel['O'][0]) if 'O' in excel else 0],amount[int(excel['P'][0]) if 'P' in excel else 0],amount[int(excel['Q'][0]) if 'Q' in excel else 0],amount[int(excel['R'][0]) if 'R' in excel else 0],amount[int(excel['S'][0]) if 'S' in excel else 0],amount[int(excel['T'][0]) if 'T' in excel else 0],amount[int(excel['U'][0]) if 'U' in excel else 0],amount[int(excel['V'][0]) if 'V' in excel else 0],amount[int(excel['W'][0]) if 'W' in excel else 0],amount[int(excel['X'][0]) if 'X' in excel else 0],amount[int(excel['Y'][0]) if 'Y' in excel else 0],amount[int(excel['Z'][0]) if 'Z' in excel else 0],amount[int(excel['AA'][0]) if 'AA' in excel else 0],amount[int(excel['AB'][0]) if 'AB' in excel else 0],amount[int(excel['AC'][0]) if 'AC' in excel else 0],amount[int(excel['AD'][0]) if 'AD' in excel else 0],amount[int(excel['AE'][0]) if 'AE' in excel else 0],amount[int(excel['AF'][0]) if 'AF' in excel else 0],amount[int(excel['AG'][0]) if 'AG' in excel else 0],amount[int(excel['AH'][0]) if 'AH' in excel else 0],amount[int(excel['AI'][0]) if 'AI' in excel else 0],amount[int(excel['AJ'][0]) if 'AJ' in excel else 0],amount[int(excel['AK'][0]) if 'AK' in excel else 0],amount[int(excel['AL'][0]) if 'AL' in excel else 0],amount[int(excel['AM'][0]) if 'AM' in excel else 0],amount[int(excel['AN'][0]) if 'AN' in excel else 0],amount[int(excel['AO'][0]) if 'AO' in excel else 0],amount[int(excel['AP'][0]) if 'AP' in excel else 0],amount[int(excel['AQ'][0]) if 'AQ' in excel else 0],amount[int(excel['AR'][0]) if 'AR' in excel else 0],amount[int(excel['AS'][0]) if 'AS' in excel else 0],amount[int(excel['AT'][0]) if 'AT' in excel else 0],amount[int(excel['AU'][0]) if 'AU' in excel else 0],amount[int(excel['AV'][0]) if 'AV' in excel else 0],amount[int(excel['AW'][0]) if 'AW' in excel else 0],amount[int(excel['AX'][0]) if 'AX' in excel else 0],amount[int(excel['AY'][0]) if 'AY' in excel else 0],amount[int(excel['AZ'][0]) if 'AZ' in excel else 0]])
                            i += 1

                        if self.middle_status == True and amount[10] == '中':
                            ws.row_dimensions[i+1].height = 25
                            ws.append([i,amount[int(excel['B'][0]) if 'B' in excel else 0],amount[int(excel['C'][0]) if 'C' in excel else 0],amount[int(excel['D'][0]) if 'D' in excel else 0],amount[int(excel['E'][0]) if 'E' in excel else 0],amount[int(excel['F'][0]) if 'F' in excel else 0],amount[int(excel['G'][0]) if 'G' in excel else 0],amount[int(excel['H'][0]) if 'H' in excel else 0],amount[int(excel['I'][0]) if 'I' in excel else 0],amount[int(excel['J'][0]) if 'J' in excel else 0],amount[int(excel['K'][0]) if 'K' in excel else 0],amount[int(excel['L'][0]) if 'L' in excel else 0],amount[int(excel['M'][0]) if 'M' in excel else 0],amount[int(excel['N'][0]) if 'N' in excel else 0],amount[int(excel['O'][0]) if 'O' in excel else 0],amount[int(excel['P'][0]) if 'P' in excel else 0],amount[int(excel['Q'][0]) if 'Q' in excel else 0],amount[int(excel['R'][0]) if 'R' in excel else 0],amount[int(excel['S'][0]) if 'S' in excel else 0],amount[int(excel['T'][0]) if 'T' in excel else 0],amount[int(excel['U'][0]) if 'U' in excel else 0],amount[int(excel['V'][0]) if 'V' in excel else 0],amount[int(excel['W'][0]) if 'W' in excel else 0],amount[int(excel['X'][0]) if 'X' in excel else 0],amount[int(excel['Y'][0]) if 'Y' in excel else 0],amount[int(excel['Z'][0]) if 'Z' in excel else 0],amount[int(excel['AA'][0]) if 'AA' in excel else 0],amount[int(excel['AB'][0]) if 'AB' in excel else 0],amount[int(excel['AC'][0]) if 'AC' in excel else 0],amount[int(excel['AD'][0]) if 'AD' in excel else 0],amount[int(excel['AE'][0]) if 'AE' in excel else 0],amount[int(excel['AF'][0]) if 'AF' in excel else 0],amount[int(excel['AG'][0]) if 'AG' in excel else 0],amount[int(excel['AH'][0]) if 'AH' in excel else 0],amount[int(excel['AI'][0]) if 'AI' in excel else 0],amount[int(excel['AJ'][0]) if 'AJ' in excel else 0],amount[int(excel['AK'][0]) if 'AK' in excel else 0],amount[int(excel['AL'][0]) if 'AL' in excel else 0],amount[int(excel['AM'][0]) if 'AM' in excel else 0],amount[int(excel['AN'][0]) if 'AN' in excel else 0],amount[int(excel['AO'][0]) if 'AO' in excel else 0],amount[int(excel['AP'][0]) if 'AP' in excel else 0],amount[int(excel['AQ'][0]) if 'AQ' in excel else 0],amount[int(excel['AR'][0]) if 'AR' in excel else 0],amount[int(excel['AS'][0]) if 'AS' in excel else 0],amount[int(excel['AT'][0]) if 'AT' in excel else 0],amount[int(excel['AU'][0]) if 'AU' in excel else 0],amount[int(excel['AV'][0]) if 'AV' in excel else 0],amount[int(excel['AW'][0]) if 'AW' in excel else 0],amount[int(excel['AX'][0]) if 'AX' in excel else 0],amount[int(excel['AY'][0]) if 'AY' in excel else 0],amount[int(excel['AZ'][0]) if 'AZ' in excel else 0]])
                            i += 1

                        if self.low_status == True and amount[10] == '低':
                            ws.row_dimensions[i+1].height = 25
                            ws.append([i,amount[int(excel['B'][0]) if 'B' in excel else 0],amount[int(excel['C'][0]) if 'C' in excel else 0],amount[int(excel['D'][0]) if 'D' in excel else 0],amount[int(excel['E'][0]) if 'E' in excel else 0],amount[int(excel['F'][0]) if 'F' in excel else 0],amount[int(excel['G'][0]) if 'G' in excel else 0],amount[int(excel['H'][0]) if 'H' in excel else 0],amount[int(excel['I'][0]) if 'I' in excel else 0],amount[int(excel['J'][0]) if 'J' in excel else 0],amount[int(excel['K'][0]) if 'K' in excel else 0],amount[int(excel['L'][0]) if 'L' in excel else 0],amount[int(excel['M'][0]) if 'M' in excel else 0],amount[int(excel['N'][0]) if 'N' in excel else 0],amount[int(excel['O'][0]) if 'O' in excel else 0],amount[int(excel['P'][0]) if 'P' in excel else 0],amount[int(excel['Q'][0]) if 'Q' in excel else 0],amount[int(excel['R'][0]) if 'R' in excel else 0],amount[int(excel['S'][0]) if 'S' in excel else 0],amount[int(excel['T'][0]) if 'T' in excel else 0],amount[int(excel['U'][0]) if 'U' in excel else 0],amount[int(excel['V'][0]) if 'V' in excel else 0],amount[int(excel['W'][0]) if 'W' in excel else 0],amount[int(excel['X'][0]) if 'X' in excel else 0],amount[int(excel['Y'][0]) if 'Y' in excel else 0],amount[int(excel['Z'][0]) if 'Z' in excel else 0],amount[int(excel['AA'][0]) if 'AA' in excel else 0],amount[int(excel['AB'][0]) if 'AB' in excel else 0],amount[int(excel['AC'][0]) if 'AC' in excel else 0],amount[int(excel['AD'][0]) if 'AD' in excel else 0],amount[int(excel['AE'][0]) if 'AE' in excel else 0],amount[int(excel['AF'][0]) if 'AF' in excel else 0],amount[int(excel['AG'][0]) if 'AG' in excel else 0],amount[int(excel['AH'][0]) if 'AH' in excel else 0],amount[int(excel['AI'][0]) if 'AI' in excel else 0],amount[int(excel['AJ'][0]) if 'AJ' in excel else 0],amount[int(excel['AK'][0]) if 'AK' in excel else 0],amount[int(excel['AL'][0]) if 'AL' in excel else 0],amount[int(excel['AM'][0]) if 'AM' in excel else 0],amount[int(excel['AN'][0]) if 'AN' in excel else 0],amount[int(excel['AO'][0]) if 'AO' in excel else 0],amount[int(excel['AP'][0]) if 'AP' in excel else 0],amount[int(excel['AQ'][0]) if 'AQ' in excel else 0],amount[int(excel['AR'][0]) if 'AR' in excel else 0],amount[int(excel['AS'][0]) if 'AS' in excel else 0],amount[int(excel['AT'][0]) if 'AT' in excel else 0],amount[int(excel['AU'][0]) if 'AU' in excel else 0],amount[int(excel['AV'][0]) if 'AV' in excel else 0],amount[int(excel['AW'][0]) if 'AW' in excel else 0],amount[int(excel['AX'][0]) if 'AX' in excel else 0],amount[int(excel['AY'][0]) if 'AY' in excel else 0],amount[int(excel['AZ'][0]) if 'AZ' in excel else 0]])
                            i += 1
                    for row in ws['A2:{}{}'.format(letter[len(excel)-1],i)]:
                        for cell in row:
                            cell.font = font
                            cell.border = border
                            cell.alignment = alignment
                    
                ws.delete_cols(len_excel,len(letter))
                wb.save(self.folder_end+'/汇总-漏洞跟踪表/高中风险漏洞跟踪表--汇总.xlsx')
                del vul_all_list[:]
                vul_all_detail.clear()

            self.log_return.emit('漏洞跟踪表导出完成，保存在输出路径 汇总-漏洞跟踪表 目录下。')
            shutil.rmtree('temp')
            endtime = datetime.datetime.now()
            self.log_return.emit('导出花时：{}秒...\n'.format((endtime - starttime).seconds))

        if self.port_status:
            try:
                shutil.rmtree(self.folder_end+'/汇总-端口对应关系表')
            except Exception as e:
                pass
            self.log_return.emit('正在导出端口，请稍后！')
            starttime = datetime.datetime.now()
            os.mkdir(self.folder_end+'/汇总-端口对应关系表')
            dirList = os.listdir(self.folder_start)
            for name in dirList:
                all_file_name = re.findall(Port_File_re().file_re,name)
                for file_name in all_file_name:
                    uzip = zipfile.ZipFile(self.folder_start+'/'+file_name)
                    i = 1
                    wb = openpyxl.Workbook()
                    wps = wb.active

                    wps.column_dimensions['A'].width = 16.5
                    wps.column_dimensions['B'].width = 16
                    wps.column_dimensions['C'].width = 20
                    wps.column_dimensions['D'].width = 30
                    wps.column_dimensions['E'].width = 25
                    wps.column_dimensions['F'].width = 28
                    wps.column_dimensions['G'].width = 42
                    wps.column_dimensions['H'].width = 17


                    wps.title = '端口数据'
                    wps['A1'] = '设备端口和服务信息表'
                    wps.merge_cells('A1:H1')
                    wps['A2'] = '收集时间'
                    wps.merge_cells('A2:B2')
                    wps.merge_cells('C2:D2')
                    wps['E2'] = '所属系统'
                    wps.merge_cells('F2:H2')
                    wps['A3'] = '填表人'
                    wps['C3'] = self.name_ini
                    wps.merge_cells('A3:B3')
                    wps.merge_cells('C3:D3')
                    wps['E3'] = '系统责任人'
                    wps.merge_cells('F3:H3')
                    wps['A4'] = 'IP地址'
                    wps['B4'] = '端口'
                    wps['C4'] = '协议'
                    wps['D4'] = '服务'
                    wps['E4'] = '状态'
                    wps['F4'] = '访问权限开放范围'
                    wps['G4'] = '应用说明'
                    wps['H4'] = '备注'

                    # 样式
                    font = Font(size=12, name='宋体')
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 对齐
                    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    #设置第一行的格式
                    one_font = Font(size=12, bold=True, name='宋体')
                    for one_style in wps['A1:H1']:
                        for one_cell in one_style:
                            one_cell.font = one_font
                            one_cell.border = border
                            one_cell.alignment = alignment

                    for excel_style in wps['A2:H4']:
                        for excel_cell in excel_style:
                            excel_cell.font = font
                            excel_cell.border = border
                            excel_cell.alignment = alignment

                    for uzip_content in uzip.namelist():
                        all_uzip_content = re.findall(Port_File_re().uzip_re,uzip_content)
                        for all_uzip in all_uzip_content:
                            htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                            vul_title = re.findall(Port_File_re().all_title_re,htmlcont_zip,re.S|re.M)
                            for title_content in vul_title:
                                pass

                            vul_host = re.findall(Port_File_re().host_re,htmlcont_zip,re.S|re.M)
                            for host_content in vul_host:
                              for vul_port in re.findall(Port_File_re().port_re,host_content[2],re.S|re.M):
                                wps.row_dimensions[i+4].height = 15
                                wps['C2'] = '%s' % host_content[1]
                                wps.append([host_content[0],vul_port[0].replace(' ','').strip(),vul_port[1].replace(' ','').strip(),vul_port[2].replace(' ','').strip(),vul_port[3].replace(' ','').strip()])
                                for row in wps['A%s:H%s'%(i+4,i+4)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                i += 1

                    wb.save(self.folder_end+'/汇总-端口对应关系表/端口服务对应关系表--%s.xlsx' % title_content)
                    self.log_return.emit('导出 %s'%title_content)

            endtime = datetime.datetime.now()
            self.log_return.emit('所有端口导出完成，保存在输出路径 汇总-端口对应关系表 目录下。')
            self.log_return.emit('导出花时：%s秒...\n'%(endtime - starttime).seconds)

        if self.web_status:
            try:
                shutil.rmtree(self.folder_end+'/汇总-WEB网站')
            except Exception as e:
                pass
            self.log_return.emit('正在WEB网站，请稍后！')
            starttime = datetime.datetime.now()
            os.mkdir(self.folder_end+'/汇总-WEB网站')
            dirList = os.listdir(self.folder_start)
            for name in dirList:
                all_file_name = re.findall(Port_File_re().file_re,name)
                for file_name in all_file_name:
                    uzip = zipfile.ZipFile(self.folder_start+'/'+file_name)

                    x = 1
                    web = openpyxl.Workbook()
                    wes = web.active
                    wes.title = 'WEB网站'

                    wes.column_dimensions['A'].width = 16.5
                    wes.column_dimensions['B'].width = 16
                    wes.column_dimensions['C'].width = 20
                    wes.column_dimensions['D'].width = 30
                    wes.column_dimensions['E'].width = 25
                    wes.column_dimensions['F'].width = 45

                    wes['A1'] = 'IP地址'
                    wes['B1'] = '端口'
                    wes['C1'] = '协议'
                    wes['D1'] = '服务'
                    wes['E1'] = '状态'
                    wes['F1'] = 'WEB网站信息'

                    # 样式
                    font = Font(size=12, name='宋体')
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 对齐
                    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    for uzip_content in uzip.namelist():
                        all_uzip_content = re.findall(Port_File_re().uzip_re,uzip_content)
                        for all_uzip in all_uzip_content:
                            htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                            vul_title = re.findall(Port_File_re().all_title_re,htmlcont_zip,re.S|re.M)
                            for title_content in vul_title:
                                pass

                            vul_host = re.findall(Port_File_re().host_re,htmlcont_zip,re.S|re.M)
                            for host_content in vul_host:
                              for vul_port in re.findall(Port_File_re().port_re,host_content[2],re.S|re.M):
                                vul_web = re.findall(Port_File_re().http_re,vul_port[2].replace(' ','').strip(),re.S|re.M)
                                if vul_web:
                                    wes.row_dimensions[x+1].height = 15
                                    wes.append([host_content[0],vul_port[0].replace(' ','').strip(),vul_port[1].replace(' ','').strip(),vul_port[2].replace(' ','').strip(),vul_port[3].replace(' ','').strip(),'http://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip()])
                                    x += 1
                                vul_web = re.findall(Port_File_re().https_re,vul_port[2].replace(' ','').strip(),re.S|re.M)
                                if vul_web:
                                    wes.row_dimensions[x+1].height = 15
                                    wes.append([host_content[0],vul_port[0].replace(' ','').strip(),vul_port[1].replace(' ','').strip(),vul_port[2].replace(' ','').strip(),vul_port[3].replace(' ','').strip(),'https://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip()])
                                    x += 1
                                vul_web = re.findall(Port_File_re().www_re,vul_port[2].replace(' ','').strip(),re.S|re.M)
                                if vul_web:
                                    wes.row_dimensions[x+1].height = 15
                                    wes.append([host_content[0],vul_port[0].replace(' ','').strip(),vul_port[1].replace(' ','').strip(),vul_port[2].replace(' ','').strip(),vul_port[3].replace(' ','').strip(),'http://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip()])
                                    x += 1

                    for row in wes['A1:F{}'.format(x)]:
                        for cell in row:
                            cell.font = font
                            cell.border = border
                            cell.alignment = alignment

                    web.save(self.folder_end+'/汇总-WEB网站/WEB网站--%s.xlsx' % title_content)
                    self.log_return.emit('导出 %s'%title_content)
            endtime = datetime.datetime.now()
            self.log_return.emit('所有WEB网站导出导出完成，保存在输出路径 汇总-WEB网站 目录下。')
            self.log_return.emit('导出花时：%s秒...'%(endtime - starttime).seconds)
